import base64
from Crypto.Cipher import AES
import os
import ray
import platform
import json
import pandas as pd
if platform.system().lower() == 'linux':
    import fcntl
elif platform.system().lower() == 'windows':
    import fcntlock as fcntl
else:
    raise Exception("系统：{} 暂不支持！".format(platform.system().lower()))
import time
import re
from openpyxl import load_workbook
from filelock import FileLock
import polars as pl

CIPHER_KEY = 'eUMMn9zWE8EBPHt6hkNooQ=='
xquant_id = None

def decrypt(ciphertext):
    def _unpad(s):
        return s[:-ord(s[len(s) - 1:])]
    CIPHER_KEY_bytes = CIPHER_KEY.encode('utf-8')
    enc = base64.b64decode(ciphertext)
    iv = enc[:AES.block_size]
    if not os.path.exists('/opt/anaconda3/lib/python3.11'):
        cipher = AES.new(CIPHER_KEY, AES.MODE_CBC, iv)
    else:
        cipher = AES.new(CIPHER_KEY_bytes, AES.MODE_CBC, iv)
    return _unpad(cipher.decrypt(enc[AES.block_size:])).decode('utf-8')

def get_user_id():
    XQUANT_CONF_FILE = os.environ['XQUANT_CONF_FILE']
    with open(XQUANT_CONF_FILE, 'r') as f:
        r = f.readlines()
        for line in r:
            if line.startswith("userId"):
                return line.strip().split("=")[-1].strip()


def start_ray_cluster(object_spill_path = None, num_cpus = None, restart = True):
    if ray.is_initialized() and restart:
        ray.shutdown()
    if not object_spill_path:
        try:
            user_id = get_user_id()
            if os.path.exists(f"/dfs/user/{user_id}/"):
                object_spill_path = f"/dfs/user/{user_id}/"
            else:
                object_spill_path = f"/data/user/{user_id}/"
        except:
            object_spill_path = "/data/user/013150/tmp"

    # print("object_spill_path:", object_spill_path)
    if not num_cpus:
        ray.init(_system_config={
            "object_spilling_config": json.dumps(
                {"type": "filesystem", "params": {"directory_path": object_spill_path}},
            )},ignore_reinit_error=True)
    else:
        ray.init(num_cpus=num_cpus, _system_config={
            "object_spilling_config": json.dumps(
                {"type": "filesystem", "params": {"directory_path": object_spill_path}},
            )}, ignore_reinit_error=True)


class FileLock2:
    def __init__(self, filename):
        self.filename = filename
        self.handle = open(self.filename, 'w')

    def acquire(self):
        fcntl.lockf(self.handle.fileno(), fcntl.LOCK_EX)# | fcntl.LOCK_NB)  # LOCK_EX与LOCK_NB按位或使用不阻塞，报OSError错误


    def release(self):
        # 延迟0.5秒释放锁，其他进程非阻塞return，不再写入文件
        time.sleep(0.5)
        if platform.system().lower() == 'linux':
            fcntl.lockf(self.handle.fileno(), fcntl.LOCK_UN)
        elif platform.system().lower() == 'windows':
            fcntl.unlock(self.handle.fileno())
        else:
            raise Exception("系统：{} 暂不支持！".format(platform.system().lower()))

    def close(self):
        try:
            self.handle.close()
        except:
            pass

    def __del__(self):
        try:
            self.handle.close()
        except:
            pass


def save_and_append_xlsx(result_df, sheet_name, overwrite_col=None, output_path="./a.xlsx"):
    result_df_s = result_df.copy()
    result_df_s["SYMBOL"] = sheet_name
    result_df_s.reset_index(inplace=True)
    # 使用文件锁确保写入安全
    lock_path = output_path + '.lock'
    with FileLock(lock_path):
        if not os.path.exists(output_path):
            result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
        else:
            try:
                book = load_workbook(output_path)
            except:
                result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
                return
            pandas_v = pd.__version__
            # pandas<1.3.0时if_sheet_exists="replace"不生效
            if pandas_v < "1.3.0":
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    try:
                        if overwrite_col:
                            try:
                                df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                            except:
                                df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                            df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                            result_df_s = pd.concat([result_df_s, df_e])
                    except:
                        pass
                    if sheet_name in writer.sheets:
                        # 清空现有工作表
                        worksheet = writer.sheets[sheet_name]
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                                       max_col=worksheet.max_column):
                            for cell in row:
                                cell.value = None
                    result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
                # 保存更改
                writer.save()

            else:
                with pd.ExcelWriter(output_path, engine='openpyxl', mode="a", if_sheet_exists="replace") as writer:
                    try:
                        if overwrite_col:
                            try:
                                df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                            except:
                                df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                            df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                            result_df_s = pd.concat([result_df_s, df_e])
                        result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        print("save_and_append_xlsx exception:", e)
                        result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)

def save_and_append_xlsx_nolock(result_df, sheet_name, overwrite_col=None, output_path="./a.xlsx"):
    result_df_s = result_df.copy()
    result_df_s["SYMBOL"] = sheet_name
    if "index" not in result_df_s.columns:
        result_df_s.reset_index(inplace=True)
    if not os.path.exists(output_path):
        print(f"{output_path}不存在，直接to_excel写入")
        result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
    else:
        try:
            book = load_workbook(output_path)
        except:
            print(f"{output_path}文件损坏，直接to_excel写入")
            result_df_s.to_excel(output_path, sheet_name=sheet_name, index=False)
            return
        pandas_v = pd.__version__
        # pandas<1.3.0时if_sheet_exists="replace"不生效
        if pandas_v < "1.3.0":
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                try:
                    if overwrite_col:
                        try:
                            df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                        except:
                            df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                        df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                        result_df_s = pd.concat([result_df_s, df_e])
                except:
                    pass
                if sheet_name in writer.sheets:
                    # 清空现有工作表
                    worksheet = writer.sheets[sheet_name]
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                                   max_col=worksheet.max_column):
                        for cell in row:
                            cell.value = None
                result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
            # 保存更改
            writer.save()

        else:
            with pd.ExcelWriter(output_path, engine='openpyxl', mode="a", if_sheet_exists="replace") as writer:
                try:
                    if overwrite_col:
                        try:
                            df_e = pd.read_excel(output_path, sheet_name=sheet_name)
                        except:
                            df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
                        df_e = df_e[~df_e[overwrite_col].isin(set(result_df_s[overwrite_col].tolist()))]
                        print(f"{output_path} 过滤历史数据，然后写入")
                        result_df_s = pd.concat([result_df_s, df_e])
                    result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    print("save_and_append_xlsx exception:", e)
                    result_df_s.to_excel(writer, sheet_name=sheet_name, index=False)

def save_and_append_xlsx_pl(result_df, sheet_name, overwrite_col=None, output_path="./a.xlsx"):
    result_df_s = result_df.copy()
    if not isinstance(result_df_s, pl.DataFrame):
        result_df_s = pl.from_pandas(result_df_s)
    # result_df_s["SYMBOL"] = sheet_name
    result_df_s = result_df_s.with_columns(pl.lit(sheet_name).alias("SYMBOL"))
    # if "index" not in result_df_s.columns:
    #     result_df_s.reset_index(inplace=True)
    print(result_df_s.head())
    print(result_df_s.columns)
    if not os.path.exists(output_path):
        # pl.write_excel(output_path, {sheet_name: result_df_s})
        result_df_s.to_excel(output_path, sheet_name=sheet_name)
    else:
        if overwrite_col:
            try:
                df_e = pl.read_excel(output_path, sheet_name=sheet_name)
                increase_data = result_df_s[overwrite_col].unique().tolist()
                df_e = df_e.filter(~pl.col(overwrite_col).isin(increase_data))
                result_df_s = pl.concat([result_df_s, df_e])
            except Exception as e:
                print(f"{output_path}：{e}")
                result_df_s.to_excel(output_path, sheet_name=sheet_name)
        result_df_s.to_excel(output_path, sheet_name=sheet_name)



save_backtest_result = save_and_append_xlsx
save_backtest_result_nolock = save_and_append_xlsx_nolock
save_backtest_result_pl = save_and_append_xlsx_pl


def save_and_append_parquet(symbol_name, signal_df, save_parquet_path, overwrite_col = "DATE"):
    """
    :param symbol_name:
    :param signal_df:
    :param save_parquet_path:
    :param overwrite_col: 需要覆盖的列， 该列的值不允许有重复，比如DATE列已有20191010的数据，需要删除后再写入
    :return:
    """
    dir_name = os.path.dirname(save_parquet_path)
    os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(save_parquet_path):
        try:
            source_signal_df = pd.read_parquet(save_parquet_path)
            source_signal_df = source_signal_df[~source_signal_df[overwrite_col].isin(set(signal_df[overwrite_col].tolist()))]
            if not "SYMBOL" in source_signal_df.columns:
                source_signal_df["SYMBOL"] = symbol_name
            new_signal_df = pd.concat([source_signal_df, signal_df], axis=0)
            new_signal_df.to_parquet(save_parquet_path)
        except:
            print("WARNING: 信号文件损坏：", save_parquet_path)
            if not "SYMBOL" in signal_df.columns:
                signal_df["SYMBOL"] = symbol_name
            signal_df.to_parquet(save_parquet_path)
    else:
        print("WARNING: 信号文件不存在：", save_parquet_path)
        if not "SYMBOL" in signal_df.columns:
            signal_df["SYMBOL"] = symbol_name
        signal_df.to_parquet(save_parquet_path)


def save_and_append_pickle(detail_data, save_path, overwrite_col="date"):
    print("文件路径: ", save_path)
    if os.path.exists(save_path):
        try:
            source_detail_df = pd.read_pickle(save_path)
            source_detail_df = source_detail_df[~source_detail_df[overwrite_col].isin(set(detail_data[overwrite_col].tolist()))]
            new_signal_df = pd.concat([source_detail_df, detail_data], axis=0)
            new_signal_df.to_pickle(save_path)
            print("文件日增数据完成")
        except Exception as e:
            print("ERROR: ", e)
            print("WARNING: 文件损坏：", save_path)
            detail_data.to_pickle(save_path)
    else:
        print("WARNING: 文件不存在：", save_path)
        detail_data.to_pickle(save_path)


def save_and_append_parquet2(signal_df, save_parquet_path, overwrite_col = "DATE"):
    """
    :param signal_df:
    :param save_parquet_path:
    :param overwrite_col: 需要覆盖的列， 该列的值不允许有重复，比如DATE列已有20191010的数据，需要删除后再写入
    :return:
    """
    dir_name = os.path.dirname(save_parquet_path)
    os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(save_parquet_path):
        try:
            source_signal_df = pd.read_parquet(save_parquet_path)
            source_signal_df = source_signal_df[~source_signal_df[overwrite_col].isin(set(signal_df[overwrite_col].tolist()))]
            new_signal_df = pd.concat([source_signal_df, signal_df], axis=0)
            new_signal_df.to_parquet(save_parquet_path)
        except:
            print("WARNING: 信号文件损坏：", save_parquet_path)
            signal_df.to_parquet(save_parquet_path)
    else:
        print("WARNING: 信号文件不存在：", save_parquet_path)
        signal_df.to_parquet(save_parquet_path)


def get_cache_data_and_lack_param(file_path, increase_col, calc_params, sheet=None):
    """
    获取已缓存的数据及需要计算的increase_col列值的列表
    :param file_path: 文件路径
    :param file_type: 文件类型，可以是parquet/xlsx/pkl
    :param increase_col: 文件新增数据的列名，比如日期，标的等
    :param calc_params: 需要计算的参数(日期或标的)列表
    :param sheet: 如果file_type=xlsx则sheet不能为空，默认None
    :return:
    """
    file_type = file_path.split(".")[-1]
    if file_type not in ["parquet", "pkl", "xlsx"]:
        raise Exception("file_path的支持的文件类型为parquet、pkl、xlsx")
    if os.path.exists(file_path):
        if file_type == "parquet":
            df_cache = pd.read_parquet(file_path)
        elif file_type == "pkl":
            df_cache = pd.read_pickle(file_path)
        elif file_type == "xlsx":
            if not sheet:
                raise Exception("【file_type】为xlsx时sheet不能为None或空")
            try:
                df_cache = pd.read_excel(file_path, sheet_name=sheet)
            except Exception as e:
                print(f"{file_path}--> {e}")
                no_cache_param = calc_params
                return pd.DataFrame(), no_cache_param
        else:
            raise Exception("file_type仅支持 parquet、pkl、xlsx类型。")
        df_cache = df_cache[df_cache[increase_col].isin(calc_params)]
        cache_param = df_cache[increase_col].tolist()
        no_cache_param = list(set(calc_params) - set(cache_param))
    else:
        df_cache = pd.DataFrame()
        no_cache_param = calc_params
    return df_cache, no_cache_param


def update_model_config(label_name, exp_name=None, version_alias=None,
                        symbol_list=None, data_type=None, is_delete=False,
                        file_path = "/dfs/group/800657/COO/packages/model_config.json"):
    """
    更新存储配置文件
    :param label_name: 标签
    :param exp_name: 实验名称
    :param version_alias: 模型别名
    :param symbol_list: 股票列表
    :param is_delete: 是否删除信息，默认否，
    :return:
    # { 'LabelFirstPeak_th10_121s': {'KG101_model': {'HS_tick2': {'symbol_list':['300750.SZ', '601318.SH'],'data_type':'tick'}}}}
    # Tips is_delete=True删除配置信息，配置信息存储为字典结构，如上所示，
    # 如果删除某个模型别名则label_name，exp_name,version_alias必传，如果exp_name只有一个模型，则exp_name也会被删除，类推标签一样
    # 如果删除标签或实验，则下级信息可不填
    """
    if os.path.exists(file_path):
        with open(file_path, "r") as fr:
            model_config = json.load(fr)
        if is_delete:
            try:
                if version_alias and model_config[label_name][exp_name].get(version_alias):
                    del model_config[label_name][exp_name][version_alias]
                    if not model_config[label_name][exp_name]:
                        del model_config[label_name][exp_name]
                    if not model_config[label_name]:
                        del model_config[label_name]
                elif exp_name and model_config[label_name].get(exp_name):
                    del model_config[label_name][exp_name]
                    if not model_config[label_name]:
                        del model_config[label_name]
                elif label_name and model_config.get(label_name):
                    del model_config[label_name]
            except Exception as e:
                print(f"删除配置信息出错：KeyError: {e} !")
        else:
            if not isinstance(symbol_list, list) or len(symbol_list) == 0:
                raise Exception("symbol_list为列表类型，且不为空。")
            if not isinstance(data_type, str):
                raise Exception("data_type为string类型，如：enhanced_tick_norm, tick_l2p")
            if label_name not in model_config:
                model_config[label_name] = {
                    exp_name: {version_alias: {"symbol_list": symbol_list, "data_type": data_type}}}
            else:
                if exp_name not in model_config[label_name]:
                    model_config[label_name][exp_name] = {
                        version_alias: {"symbol_list": symbol_list, "data_type": data_type}}
                else:
                    model_config[label_name][exp_name][version_alias] = {"symbol_list": symbol_list,
                                                                         "data_type": data_type}
        with open(file_path, "w") as fw:
            json.dump(model_config, fw)
    else:
        if not isinstance(symbol_list, list) or len(symbol_list) == 0:
            raise Exception("symbol_list为列表类型，且不为空。")
        if not isinstance(data_type, str):
            raise Exception("data_type为string类型，如：enhanced_tick_norm, tick_l2p")
        model_config = {label_name: {exp_name: {version_alias: {"symbol_list": symbol_list, "data_type": data_type}}}}
        with open(file_path, "w") as f:
            json.dump(model_config, f)


def load_model_config(file_path = "/dfs/group/800657/COO/packages/model_config.json"):
    # 返回列表元素为元组的结构，即：[(),()]
    model_config = []
    if os.path.exists(file_path):
        with open(file_path, "r") as fr:
            model_config_dict = json.load(fr)
        for label in model_config_dict:
            for exp in model_config_dict[label]:
                for alias in model_config_dict[label][exp]:
                    model_config.append((label, exp, alias, model_config_dict[label][exp][alias]["symbol_list"],
                                         model_config_dict[label][exp][alias]["data_type"]))
    else:
        print(f"暂未缓存配置文件：{file_path}")
    return model_config


def save_winloss_data(res_df, save_parquet_path, sub_cols):
    """

    :param res_df: 需要存储的DataFrame
    :param save_parquet_path: 存储路径，parquet文件
    :param sub_cols: 可以确定一条数据唯一性的列组合，列表类型，用于去重
    :return:
    """
    dir_name = os.path.dirname(save_parquet_path)
    os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(save_parquet_path):
        try:
            cache_df = pd.read_parquet(save_parquet_path)
            new_df = pd.concat([cache_df, res_df])
            new_df = new_df.drop_duplicates(subset=sub_cols, keep="last")
            new_df.to_parquet(save_parquet_path)
        except:
            print(f"WARNING: {save_parquet_path}文件损坏，to_parquet写入。")
            res_df.to_parquet(save_parquet_path)
    else:
        print(f"WARNING: {save_parquet_path} 文件不存在, to_parquet写入。")
        res_df.to_parquet(save_parquet_path)

def save_winloss_data_xlsx(result_df, sign_col, output_path="./a.xlsx", overwrite_col=None):
    """
    存储止盈止损数据到Excel
    :param result_df: DataFrame，止盈止损汇总数据
    :param sign_col: 标识字段名，筛选result_df，sign_col字段的值会作为一个sheet_name
    :param output_path: 存储路径
    :param overwrite_col: 覆盖字段，默认None，如为string，则会用该字段筛选已缓存数据不在result_df中的部分
    :return:
    """
    result_df_s = result_df.copy()
    data_keys = list(set(result_df_s[sign_col].tolist()))
    if not os.path.exists(output_path):
        print(f"{output_path}不存在，直接to_excel写入")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for key in data_keys:
                res_df_p = result_df_s[result_df_s[sign_col] == key]
                res_df_p.to_excel(writer, sheet_name=key, index=False)
    else:
        try:
            book = load_workbook(output_path)
        except:
            print(f"{output_path}文件损坏，直接to_excel写入")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for key in data_keys:
                    res_df_p = result_df_s[result_df_s[sign_col] == key]
                    res_df_p.to_excel(writer, sheet_name=key, index=False)
                try:
                    writer.save()
                except:
                    pass
            return
        pandas_v = pd.__version__
        # pandas<1.3.0时if_sheet_exists="replace"不生效
        if pandas_v < "1.3.0":
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                for key in data_keys:
                    res_df_p = result_df_s[result_df_s[sign_col] == key]
                    if key in writer.sheets:
                        # TODO overwrite_col非None时，排序
                        if overwrite_col:
                            df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=key)
                            df_e = df_e[~df_e[overwrite_col].isin(res_df_p[overwrite_col].tolist())]
                            res_df_p = pd.concat([res_df_p, df_e])
                            res_df_p = res_df_p.sort_values(by=overwrite_col)
                        # 清空现有工作表
                        worksheet = writer.sheets[key]
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                                       max_col=worksheet.max_column):
                            for cell in row:
                                cell.value = None
                    res_df_p.to_excel(writer, sheet_name=key, index=False)
                writer.save()
        else:
            with pd.ExcelWriter(output_path, engine='openpyxl', mode="a", if_sheet_exists="replace") as writer:
                for key in data_keys:
                    res_df_p = result_df_s[result_df_s[sign_col] == key]
                    try:
                        df_e = pd.read_excel(output_path, engine='openpyxl', sheet_name=key)
                        if overwrite_col:
                            df_e = df_e[~df_e[overwrite_col].isin(res_df_p[overwrite_col].tolist())]
                            res_df_p = pd.concat([res_df_p, df_e])
                            res_df_p = res_df_p.sort_values(by=overwrite_col)
                    except Exception as e:
                        print("save_and_append_xlsx exception:", e)
                    res_df_p.to_excel(writer, sheet_name=key, index=False)



