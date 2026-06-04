# -*- coding: utf-8 -*-
"""
to output a spreadsheet for a better look

by 011478
"""

import openpyxl
import copy
import datetime
import numpy as np
import pandas as pd
from xquant.xqutils.xqfile import HDFSFile
from xquant.factordata import FactorData

fd = FactorData()

def OutputSpreadSheet(detailedOrders, result, dirPath, funStr, symbol):
    wb = openpyxl.Workbook()
    OutputDetailedOrders(wb, detailedOrders, funStr)
    orders_data = OutputResultOrders(wb, result)
    OutputDate(wb, orders_data, funStr, result, symbol)
    OutputSummary(wb, result)
    hf = HDFSFile()
    if hf.exists(dirPath):
        hf.delete(dirPath)
    with hf.open(dirPath, 'wb') as f:
         wb.save(f)


def OutputDetailedOrders(wb, detailedOrders, funStr):
    DetailedOrdersTitleList = ['tradeNo', 'code', 'orderTime', 'direction', 'price', 'quantity', 'avgPrice', 'cumQty',
                               'status', 'orderAmount', 'cumAmount']
    ws = wb.create_sheet('detailedOrders')
    if funStr is None:
        for i in range(1, len(DetailedOrdersTitleList) + 1):
            ws.cell(row=1, column=i, value=DetailedOrdersTitleList[i - 1])

        row = 2
        for i in range(len(detailedOrders)):
            for j in range(len(detailedOrders[i])):
                for k in range(1, len(DetailedOrdersTitleList) + 1):
                    if k == 1:
                        ws.cell(row=row, column=k, value = i + 1)
                    else:
                        ws.cell(row=row, column=k, value = detailedOrders[i][j].get(DetailedOrdersTitleList[k - 1]))
                row += 1


def OutputResultOrders(wb, result):
    ResultOrdersTitleList = ['code', 'date', 'startTime', 'endTime', 'holdTime',
                             'direction', 'startPrice', 'endPrice',
                             'orderAmount', 'cumAmount', 'returnRate', 'afterCostProfit']
    orders = result.get('order')
    ws = wb.create_sheet('orders')
    for i in range(1, len(ResultOrdersTitleList) + 1):
        ws.cell(row=1, column=i, value=ResultOrdersTitleList[i - 1])

    orders_data = []
    row = 2
    for i in range(len(orders)):
        startTime = datetime.datetime.strptime(orders[i].get('startTime'), '%m/%d/%Y-%H:%M:%S:%f')
        endTime = datetime.datetime.strptime(orders[i].get('endTime'), '%m/%d/%Y-%H:%M:%S:%f')
        date = str(startTime).split(" ")[0]

        orders_data_daily = [orders[i].get('code'), date, str(startTime).split(" ")[1], str(endTime).split(" ")[1],
                             str(endTime - startTime), orders[i].get('direction'), orders[i].get('startPrice'),
                             orders[i].get('endPrice'), orders[i].get('orderAmount'), orders[i].get('cumAmount'),
                             orders[i].get('returnRate') - 0.0012, orders[i].get('afterCostProfit')]
        for k in range(1, len(orders_data_daily) + 1):
            ws.cell(row=row, column=k, value=orders_data_daily[k - 1])
        row += 1
        orders_data.append(orders_data_daily)
    if len(orders_data) > 0:
        orders_data = pd.DataFrame(orders_data)
        orders_data.columns = ResultOrdersTitleList
    return orders_data


def OutputDate(wb, orders_data, funStr, result, symbol):
    DateTitleList = ['日期', '总市值', '总盈利', '交易次数', '获利次数', '胜率', '平均收益率', '利用率',
                     '交易总市值', '市值收益率', '获利收益率', '亏损收益率', '盈亏比',
                     '最大单笔亏损', '平均持仓时间', '撤单比']
    stockSize = result['initQty']                
    ws = wb.create_sheet('dailyInfo')
    
    if funStr is None:
        for i in range(1, len(DateTitleList) + 1):
            ws.cell(row=1, column=i, value=DateTitleList[i - 1])

        if len(orders_data) == 0:
            return

        dateList = sorted(np.unique(orders_data["date"]))
        
        closePrice = fd.get_factor_value("Basic_factor", [symbol], list(map(lambda x: "".join(x.split("-")), dateList)), ["close"])
        stockSize = (stockSize * closePrice["close"]).tolist()

        row = 2
        for i in range(len(dateList)):
            date = dateList[i]
            daily_data = orders_data[orders_data["date"] == date]

            afterCostProfit = sum(daily_data["afterCostProfit"])
            triggerTimes = daily_data.shape[0]
            winTimes = sum(daily_data["afterCostProfit"] > 0)
            winRate = winTimes / triggerTimes
            dailyOpenAmount = sum(daily_data["cumAmount"])

            afterCostProfit_earn = sum(daily_data[daily_data["afterCostProfit"] > 0]["afterCostProfit"])
            afterCostProfit_loss = sum(daily_data[daily_data["afterCostProfit"] < 0]["afterCostProfit"])
            dailyOpenAmount_earn = sum(daily_data[daily_data["afterCostProfit"] > 0]["cumAmount"])
            dailyOpenAmount_loss = sum(daily_data[daily_data["afterCostProfit"] < 0]["cumAmount"])
            earnReturnRate = afterCostProfit_earn / dailyOpenAmount_earn if dailyOpenAmount_earn != 0 else 0
            lossReturnRate = afterCostProfit_loss / dailyOpenAmount_loss if dailyOpenAmount_loss != 0 else 0
            EarnLossRate = round(-afterCostProfit_earn / afterCostProfit_loss, 2) if afterCostProfit_loss != 0 else "nan"

            holdTime = list(daily_data["holdTime"])
            for m in range(len(holdTime)):
                holdTime[m] = t2m(holdTime[m])
            aveHoldTime = np.mean(holdTime)
            aveHoldTime = m2t(aveHoldTime)

            dailyInfo_data = [date, stockSize[i], afterCostProfit, triggerTimes, winTimes,
                              str(int(winRate * 100)) + "%",
                              round(afterCostProfit / dailyOpenAmount * 1000, 2),
                              str(int(dailyOpenAmount / stockSize[i] * 100)) + "%", dailyOpenAmount,
                              round(afterCostProfit / stockSize[i] * 1000, 2),
                              round(earnReturnRate * 1000, 2), round(lossReturnRate * 1000, 2), EarnLossRate,
                              round(min(0, min(daily_data["returnRate"])) * 1000, 2), aveHoldTime, result['dailyCancelledRatio'][date]]
            for k in range(1, len(dailyInfo_data) + 1):
                ws.cell(row=row, column=k, value=str(dailyInfo_data[k - 1]))
            row += 1


def OutputSummary(wb, result):
    DateTitleList = ['date', 'preCostDailyProfit', 'afterCostDailyProfit', 'dailyOpenAmount', 'dailyCancelledRatio']
    ws = wb.create_sheet('summary')
    tempResult = copy.deepcopy(result)
    tempResult.pop('order', None)
    for i in range(len(DateTitleList)):
        tempResult.pop(DateTitleList[i], None)

    row = 1
    title = sorted(tempResult.keys())
    for i in range(1, len(title) + 1):
        ws.cell(row=row, column=i, value=title[i - 1])

    row += 1
    for i in range(1, len(title) + 1):
        ws.cell(row=row, column=i, value=tempResult.get(title[i - 1]))


def t2m(t):
    h, m, s = t.strip().split(".")[0].split(":")
    return int(h) * 60 + int(m) + int(s) / 60


def m2t(t):
    m = int(t)
    s = int((t - m) * 60)
    h = int(m / 60)
    m = m - h * 60
    ss = str(s)
    mm = str(m)
    if len(ss) == 1:
        ss = "0" + ss
    if len(mm) == 1:
        mm = "0" + mm
    return str(h) + ":" + mm + ":" + ss
