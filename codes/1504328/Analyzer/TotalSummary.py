# -*- coding: utf-8 -*-
import openpyxl
import os
import shutil
from CONFIG import USER_ID


BT_RESULT_SOURCE_PATH = "/data/user/{}/BT_Results/Stock/sources/".format(USER_ID)
BT_RESULT_RESULT_PATH = "/data/user/{}/BT_Results/Stock/results/".format(USER_ID)


def summary(today_str, files_dir, portfolio, excel_name, config):
    absolutePath = BT_RESULT_SOURCE_PATH + today_str + '/'
    files_dir = files_dir
    path = absolutePath + files_dir + '/'
    if os.path.exists(path + 'TEMP'):
        shutil.rmtree(path + 'TEMP')
    fileList = os.listdir(path)
    try:
        fileList.remove('TEMP')
    except:
        pass
    daypath = BT_RESULT_RESULT_PATH + today_str + '/'
    daypath = daypath + files_dir
    code = config['Quantity'].keys()

    if not os.path.exists(daypath):
        os.makedirs(daypath)
    # 1. Scan the result.xls in all the directories, and record all the titles
    titleDict = {}  # key = title, value = default 0
    titleList = []  # record the title strings
    errorDict = {}  # record the exception. make sure to print once
    for i in range(len(fileList)):
        innerDir = fileList[i]
        path2 = absolutePath + files_dir + '/' + innerDir + '/'
        stock_code = innerDir
        if not stock_code in code:
            continue
        file = path + innerDir + '/' + excel_name
        if os.path.exists(file):
            shutil.copy(file, daypath + '/' + stock_code + '.xlsx')

            try:
                wb = openpyxl.load_workbook(file)
            except Exception as e:
                if innerDir not in errorDict:
                    errorDict.update({innerDir: e})
            else:
                ws = wb.get_sheet_by_name('summary')
                row = [cell.value for cell in list(ws.rows)[0]]
                for j in range(len(row)):
                    if row[j] not in titleDict:
                        titleDict.update({row[j]: 0})
                        titleList.append(row[j])

    # 2. sort the titleList and set the index to the titleDict
    titleList = sorted(titleList)
    for i in range(1, len(titleList) + 1):
        titleDict.update({titleList[i - 1]: i})

    # 3. write the first title row in TotalSummary.xls
    wbSum = openpyxl.Workbook()
    wsSum = wbSum.create_sheet('TotalSummary')
    wsSum.cell(row=1, column=1, value='ModelFileName')
    for i in range(1, len(titleList) + 1):
        wsSum.cell(row=1, column=i + 1, value=titleList[i - 1])

    # 4. read each directory again and write the corresponding summary
    row_wt = 2
    for i in range(len(fileList)):
        innerDir = fileList[i]
        stock_code = innerDir
        if not stock_code in code:
            continue
        file = path + innerDir + '/' + excel_name
        if os.path.exists(file):
            try:
                wb = openpyxl.load_workbook(file)
            except Exception as e:
                if innerDir not in errorDict:
                    errorDict.update({innerDir: e})
            else:
                ws = wb.get_sheet_by_name('summary')
                row_title = [cell.value for cell in list(ws.rows)[0]]
                row_data = [cell.value for cell in list(ws.rows)[1]]
                for j in range(len(row_title)):
                    title = row_title[j]
                    titleIndex = titleDict.get(title) + 1  # +1 because the first column is 'ModelFileName'
                    wsSum.cell(row=row_wt, column=titleIndex, value=row_data[j])
                wsSum.cell(row=row_wt, column=1, value=innerDir)
                row_wt += 1

    # finally output the error messages
    for e in errorDict.values():
        print(e)
    # write the excel file
    wbSum.save(daypath + '/TotalSummary.xlsx')


