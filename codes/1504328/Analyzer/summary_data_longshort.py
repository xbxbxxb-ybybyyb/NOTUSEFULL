import os
import re
import openpyxl
import numpy as np
import pandas as pd
from CONFIG import USER_ID


BT_RESULT_RESULT_PATH = "/data/user/{}/BT_Results/Stock/results/".format(USER_ID)


def OutputSpreadSheet(absolutePath, code):

    orders_data, stockSize = CombineResultOrders(absolutePath, code)

    wb = openpyxl.Workbook()
    OutputDate(wb, orders_data, stockSize, direction=["long "])
    wb.save(absolutePath + "/result_daily_long.xlsx")

    wb = openpyxl.Workbook()
    OutputDate(wb, orders_data, stockSize, direction=["short"])
    wb.save(absolutePath + "/result_daily_short.xlsx")


def CombineResultOrders(absolutePath, code):
    orders_data = []
    title = []
    stockSize = 0.0
    for stock_code in code:
        workbook = openpyxl.load_workbook(absolutePath + stock_code + ".xlsx")
        booksheet = workbook.get_sheet_by_name('orders')
        rows_list = list(booksheet.rows)
        title = [cell.value for cell in rows_list[0]]
        n_data = len(rows_list)
        for i in range(1, n_data):
            data = [cell.value for cell in rows_list[i]]
            orders_data.append(data)
        daily_sheet = workbook.get_sheet_by_name('summary')
        col1 = [cell.value for cell in list(daily_sheet.columns)[12]]
        if len(col1) >= 1:
            stockSize = stockSize + col1[1]
    orders_data = pd.DataFrame(orders_data)
    orders_data.columns = title
    return orders_data,  stockSize

def OutputDate(wb, orders_data, stockSize, direction=["long ", "short"]):
    DateTitleList = ['日期', '总市值', '总盈利', '交易次数', '获利次数', '胜率', '平均收益率', '利用率',
                     '交易总市值', '市值收益率', '获利收益率', '亏损收益率', '盈亏比',
                     '最大单笔亏损', '平均持仓时间']
    ws = wb.create_sheet('dailyInfo')
    for i in range(1, len(DateTitleList) + 1):
        ws.cell(row=1, column=i, value=DateTitleList[i - 1])
    dateList = sorted(np.unique(orders_data["date"]))

    row = 2
    for i in range(len(dateList)):
        date = dateList[i]
        daily_data = orders_data[orders_data["date"] == date]
        daily_data = daily_data[daily_data["direction"].isin(direction)]
        if daily_data.empty:
            continue

        afterCostProfit = sum(daily_data["afterCostProfit"])
        triggerTimes = daily_data.shape[0]
        winTimes = sum(daily_data["afterCostProfit"] > 0)
        winRate = winTimes / triggerTimes if triggerTimes != 0 else "nan"
        dailyOpenAmount = sum(daily_data["cumAmount"])

        afterCostProfit_earn = sum(daily_data[daily_data["afterCostProfit"] > 0]["afterCostProfit"])
        afterCostProfit_loss = sum(daily_data[daily_data["afterCostProfit"] < 0]["afterCostProfit"])
        dailyOpenAmount_earn = sum(daily_data[daily_data["afterCostProfit"] > 0]["cumAmount"])
        dailyOpenAmount_loss = sum(daily_data[daily_data["afterCostProfit"] < 0]["cumAmount"])
        earnReturnRate = afterCostProfit_earn / dailyOpenAmount_earn if dailyOpenAmount_earn != 0 else 0
        lossReturnRate = afterCostProfit_loss / dailyOpenAmount_loss if dailyOpenAmount_loss != 0 else 0
        EarnLossRate = round(-afterCostProfit_earn / afterCostProfit_loss, 2) if afterCostProfit_loss != 0 else "nan"

        holdTime = list(daily_data["holdTime"])
        for m in range(len(holdTime)):
            holdTime[m] = t2m(holdTime[m])
        aveHoldTime = np.mean(holdTime)
        aveHoldTime = m2t(aveHoldTime)

        dailyInfo_data = [date, stockSize, afterCostProfit, triggerTimes, winTimes,
                          str(int(winRate * 100)) + "%",
                          round(afterCostProfit / dailyOpenAmount * 1000, 2),
                          str(int(dailyOpenAmount / stockSize * 100)) + "%", dailyOpenAmount,
                          round(afterCostProfit / stockSize * 1000, 2),
                          round(earnReturnRate * 1000, 2), round(lossReturnRate * 1000, 2), EarnLossRate,
                          round(min(0, min(daily_data["returnRate"])) * 1000, 2), aveHoldTime]
        for k in range(1, len(dailyInfo_data) + 1):
            ws.cell(row=row, column=k, value=dailyInfo_data[k - 1])
        row += 1


def t2m(t):
    h, m, s = t.strip().split(".")[0].split(":")
    return int(h) * 60 + int(m) + int(s) / 60


def m2t(t):
    m = int(t)
    s = int((t - m) * 60)
    h = int(m / 60)
    m = m - h * 60
    ss = str(s)
    mm = str(m)
    if len(ss) == 1:
        ss = "0" + ss
    if len(mm) == 1:
        mm = "0" + mm
    return str(h) + ":" + mm + ":" + ss


def summary(today_str, files_dir):
    absolutePath = BT_RESULT_RESULT_PATH + today_str + '/' + files_dir + '/'
    code = []
    results = os.listdir(absolutePath) 
    for rst in results:
        if re.match(r"^\d{6}.\D{2}.*", rst):
            code.append(rst.split(".xlsx")[0])
    OutputSpreadSheet(absolutePath, code)
