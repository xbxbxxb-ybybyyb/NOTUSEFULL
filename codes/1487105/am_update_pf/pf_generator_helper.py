import math
import os
import sys
import numpy as np
import pandas as pd
import config_path
sys.path.insert(0, config_path.tools_path)
sys.path.insert(0, 'am_update_pf/')
#from optimize import optimize_hf
import optimize_hf_industry as optInd
import optimize_hf_industry_dealST_LF as optLF
import optimize_hf_industry_dealST_LF_300 as optLF300

import warnings
warnings.filterwarnings("ignore")
from scipy.stats import norm
from xquant.factordata import FactorData
import copy
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from xquant.futuredata import FutureData
fd = FutureData()

from xquant.xqutils.helper import link
lm = link.LinkMessage()

factor_tool = FactorData()

from xquant.thirdpartydata.marketdata import MarketData
ma = MarketData()
model_name_map = config_path.model_name_map
name_map = config_path.name_map
update_params = config_path.update_params
pool_stock_path = config_path.pool_stock_path

def FutureBias(today,pf_code):
    result1 = factor_tool.tradingday(today,-2)
    pre_date = result1[-2]
    date = result1[-1]

    benchmark = update_params[pf_code]['benchmark']
    contact = update_params[pf_code]['contact']
    add_capital = update_params[pf_code]['add_capital']
    name = update_params[pf_code]['name']
    open_capital = update_params[pf_code]['open_capital']        

    if benchmark == '500':
        point_price = 200
        benchmark_name = 'zz500'
        zz500_w = pd.read_pickle(config_path.basic_data_path + 'daily/ZZ500_data.pkl').loc[pre_date]
    elif benchmark == '300':
        point_price = 300
        benchmark_name = 'hs300'
        zz500_w = pd.read_pickle(config_path.basic_data_path + 'daily/HS300_data.pkl').loc[pre_date]
    
    df1 = fd.get_instrument_info(contact)
    contact_ = df1['windcode'].values[0]
    df = ma.getMDSecurityTickDataFrame(contact_,"%s140000" % pre_date,"%s150000" % pre_date,0)
    IC05 = df['LastPx'].iloc[-1]
    zz500_w = zz500_w[zz500_w>0]
    zz500_w.index=[c[:6] for c in zz500_w.index]
    # index_price = pd.read_pickle(config_path.basic_data_path + 'daily/close_000905SH.pkl').loc[pre_date].iloc[0]
    info_dict = {}
    info_dict[name] = (open_capital,pf_code)

    stats=pd.DataFrame(index=pd.MultiIndex.from_product(([date],info_dict.keys())),columns=['合约名称','新增期货手数','期货合约市值（万）','现货中属于指数成份股的市值（万）'])
    for k,v in info_dict.items():
        real_w = pd.read_excel(config_path.open_path + '%s/' % pf_code +v[1]+'_%s_' % benchmark_name +date+'_percentage_am.xls',header=1,converters={'证券代码':str})
        real_w.set_index('证券代码',inplace=True)
        real_w = real_w['设置比例']/100
        print(k,'weight sum：',real_w.sum())
        stats.loc[(date,k),'合约名称'] = contact#[:2]
        stats.loc[(date,k),'新增期货手数']= math.ceil(add_capital/point_price/IC05)
        stats.loc[(date,k),'期货合约市值（万）'] = round(stats.loc[(date,k),'新增期货手数']*point_price*IC05/1e4,2)
        stats.loc[(date,k),'现货中属于指数成份股的市值（万）'] = round((real_w.loc[zz500_w.index]*add_capital).sum()/1e4,2)
        stats.loc[(date,k),'期现偏离金额（万）'] = round(stats.loc[(date,k),'期货合约市值（万）']-stats.loc[(date,k),'现货中属于指数成份股的市值（万）'],2)
    print(stats)
    return stats

def get_act_stat(today,transaction_time,benchmark='500'):
    factor_tool = FactorData()

    # today = '20200408'

    # transaction_time = '0930vwap'

    today_dt = today
    result1 = factor_tool.tradingday(today_dt,-3)
    last_dt = result1[-2]    
    last_last_dt = result1[-3]    

    weight_path = config_path.weight_path + 'benchmark_'+benchmark+'/'
    predict_path = config_path.act_path
    all_name = 'All_'+benchmark
    weight_pm_sim = pd.read_pickle(weight_path+'simulation_w_pm.pkl')
    weight_am_sim = pd.read_pickle(weight_path+'simulation_w_am.pkl')
    weight_vwap_sim = pd.read_pickle(weight_path+'simulation_w_vwap.pkl')    
    weight_pm_real = pd.read_pickle(weight_path+'real_w_pm.pkl')
    weight_am_real = pd.read_pickle(weight_path+'real_w_am.pkl')
    weight_vwap_real = pd.read_pickle(weight_path+'real_w_vwap.pkl')

    # act stat
    if transaction_time == '1300':
        act = pd.read_csv(predict_path+'pm/'+all_name+'/%s.csv' % today_dt,header=None)
        act_before = pd.read_csv(predict_path+'am/'+all_name+'/%s.csv' % last_dt,header=None)
        w_now = weight_pm_real.loc[today_dt]
        w_before = weight_am_real.loc[today_dt]
    elif transaction_time == '0930':
        act = pd.read_csv(predict_path+'am/'+all_name+'/%s.csv' % last_dt,header=None)
        act_before = pd.read_csv(predict_path+'pm/'+all_name+'/%s.csv' % last_dt,header=None)
        w_now = weight_am_real.loc[today_dt]
        w_before = weight_pm_sim.loc[last_dt]        
    elif transaction_time == '0930vwap':
        act = pd.read_csv(predict_path+'vwap/'+all_name+'/%s.csv' % last_dt,header=None)
        act_before = pd.read_csv(predict_path+'vwap/'+all_name+'/%s.csv' % last_last_dt,header=None)
        w_now = weight_vwap_real.loc[today_dt]
        w_before = weight_vwap_sim.loc[last_dt]                
    elif transaction_time == '0930lf':
        weight_lf_sim = pd.read_pickle(weight_path+'simulation_w_lf.pkl')
        weight_lf_real = pd.read_pickle(weight_path+'real_w_lf.pkl')        
        act = pd.read_csv(predict_path+'lf/'+all_name+'/%s.csv' % last_dt,header=None)
        act_before = pd.read_csv(predict_path+'lf/'+all_name+'/%s.csv' % last_last_dt,header=None)
        w_now = weight_lf_real.loc[today_dt]
        w_before = weight_lf_sim.loc[last_dt]  
    else:
        act = pd.read_csv(predict_path+'vwap300/'+all_name+'/%s.csv' % last_dt,header=None)
        act_before = pd.read_csv(predict_path+'vwap300/'+all_name+'/%s.csv' % last_last_dt,header=None)
        w_now = weight_vwap_real.loc[today_dt]
        w_before = weight_vwap_sim.loc[last_dt]                        

    merged = act.merge(act_before,on=[0])[['1_x','1_y']]

    act_corr = merged.corr().values[0,1]

    act_stat = act[1].describe()

    act_stat.loc['corr_with_before'] = act_corr

    act_stat = round(act_stat,2)
    act_stat = act_stat.astype('str')
    act_stat_ = act_stat.to_dict()

    w_corr = np.corrcoef(w_now.values,w_before.values)[0,1]

    act_stat_['w_corr'] = str(round(w_corr,2))
    act_dict_str = ''
    for key, value in act_stat_.items():
        act_dict_str=act_dict_str+key+':'+value+', '
    return act_dict_str+', SubModel Act Status:OK'



def get_index_price(time,date):
    if time[:4] == '0930':
        start_date = date
        end_date = date
        data_api = FactorData()

        index_industry_dict = {
                        'index_code': [
                            '000001.SH',    # SZZZ
                            '000016.SH',    # SZ50
                            '000300.SH',    # HS300
                            '000905.SH',    # ZZ500
                            '000906.SH',    # ZZ800
                            '000985.CSI',    # ZZQZ
                            '000852.SH',    # ZZ1000
                            '399001.SZ',    # SZCZ
                            '399005.SZ',    # ZXBZ
                            '399006.SZ'     # CYBZ
                        ],

                        'index_factor': ['close', 'open'],
                        'index_factor_map_wind': ['S_DQ_CLOSE', 'S_DQ_OPEN'],

                        'index_data': ['HS300', 'ZZ500',  'SZ50'], # 'ZZ800',

                        'industry': [            
                            'sw1',
                            'sw2',
                            'citics1',
                            'citics2']
                    }

        index_code_list = index_industry_dict['index_code']
        index_factor_list = index_industry_dict['index_factor']
        index_factor_map_wind = index_industry_dict['index_factor_map_wind']
        index_data_list = index_industry_dict['index_data']
        industry_list = index_industry_dict['industry']



        close_df = data_api.get_factor_value('WIND_AIndexEODPrices', factors=['S_INFO_WINDCODE', 'TRADE_DT']+index_factor_map_wind, 
                                 TRADE_DT=['>='+start_date, '<='+end_date], S_INFO_WINDCODE=index_code_list).sort_values(by=['TRADE_DT'],ascending=True)

        close_df.index = close_df['S_INFO_WINDCODE'].values
    else:
        ma = MarketData()
        indexs = ['000905.SH','000300.SH','000016.SH']
        close_list = []
        for index in indexs:

            t = ma.getMDSecurityTickDataFrame(index,'%s110000' % date,'%s120000' % date)['LastPx'].iloc[-1]
            close_list.append(t)

        close_df = pd.DataFrame(close_list,index=indexs,columns=['S_DQ_CLOSE'])        
    
    return close_df


def Open_file_save(today_open_fw,time,pf_code,today_dt,benchmark):
    template_path = config_path.task_path + 'Noon_update_process/template.xlsx'
    if time[:4] == '0930':
        day = 'am'
    else:
        day = 'pm'
    tmp_path = config_path.task_path + 'Noon_update_process/temp.xlsx'
    if benchmark == '500':
        output_path = config_path.open_path + '%s/%s_zz500_%s_percentage_%s.xls'%(pf_code,pf_code,today_dt,day)
    else:
        output_path = config_path.open_path + '%s/%s_hs300_%s_percentage_%s.xls'%(pf_code,pf_code,today_dt,day)        
    shutil.copyfile(template_path, tmp_path)
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb['Sheet1']
    ws.title = 'Portfolio Details'


    source = pd.DataFrame(np.nan,columns=['code','Name', 'Market', 'Number', 'Capital', 'Capital_Weight', 'Weight'],index=today_open_fw.index)

    source['code'] = today_open_fw['证券代码']
    source['Name'] = today_open_fw['证券名称']
    source['Market'] = today_open_fw['市场名称']
    source['Weight'] = today_open_fw['权重']

    rows = dataframe_to_rows(source, header=False,index=False)

    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx+3, column=c_idx+1, value=value)
    wb.save(tmp_path)

    workbook = openpyxl.load_workbook(tmp_path)
    #outfile = f"{output_path.split('.')[0]}.xls"
    workbook.save(output_path)
    print('Save Done')
    
def act_check(act1):
    assert (act1.T).describe().loc['min'].values[0]!=(act1.T).describe().loc['max'].values[0],'act values are the same'
    assert (np.isnan(act1).T).sum().values[0] == 0,'NaN in act'
def map_act(df):
    from scipy.stats import norm
    df = round(df,10)
    df = df.rank(pct=True,axis=1)
    df[pd.DataFrame(index=df.index,columns=df.columns,data=(df.values==1))] = 1 -1/3676
    df[pd.DataFrame(index=df.index,columns=df.columns,data=(df.values==0))] = 1/3676
    df_norm = pd.DataFrame(norm.ppf(df.values),index=df.index,columns=df.columns)
    return df_norm
def norm_ppf(factor_dict):
    com = None
    for f in factor_dict:
        df = map_act(factor_dict[f])
        if com is None:
            com = df
        else:
            com += df
    return com/len(factor_dict)

def get_act(dt,model_list,excess_type,all_name,mode=0):
    predict_path = config_path.act_path
    act_list = [ ]
    for m in model_list:
        predict = pd.read_csv(predict_path+excess_type+'/'+m+'//%s.csv' % dt,header=None)
        print(predict_path+excess_type+'/'+m+'//%s.csv')
        act = pd.DataFrame(predict[1].values.reshape(1,-1),columns=predict[0].values.tolist())
        act_check(act)
        act_list.append(act)
    act_mean = norm_ppf(dict(zip(model_list,act_list)))
    if mode == 1:
        act_mean = map_act(act_mean)
    try:
        os.makedirs(predict_path+excess_type+'/'+all_name)
    except:
        pass
    (act_mean.T).to_csv(predict_path+excess_type+'/'+all_name+'/%s.csv' % dt,header=None)
    print('act_path:',predict_path+excess_type+'/'+all_name+'/%s.csv' % dt)

def load_pf(pf_code, file_path, add_idx_suffix=True):
    df = pd.read_excel(file_path, converters={'组合名称':str, '证券代码':str})
    pf = df[df['组合名称']==pf_code]
    pf.set_index(['证券代码'], inplace=True)
    # pf['市值'] = [float(v.replace(',', '')) for v in pf['市值']]
    # pf['市值'] = pf['市值'].apply(lambda x: float(x.replace(',','')))
    short_ = pf[pf['证券类别']=='股指期货']
    short_.loc[short_['持仓多空标志']=='多仓', ['持仓', '市值']] *= -1
    long_ = pf[pf['证券类别']=='股票']
    if add_idx_suffix:
        long_.index = [i + '.SH' if i[0]=='6' else i + '.SZ' for i in long_.index]
    return pf, long_, short_

def save_xlsx(file_path, sheet_dict):
    writer = pd.ExcelWriter(file_path)
    for k in sheet_dict.keys():
        sheet_dict[k].to_excel(writer, sheet_name=k)
    writer.save()

def exclude_invalid(position, valid_stk):
    pos_valid_stk = list(set(position.index).intersection(valid_stk))
    pos_valid_stk.sort()
    return position[pos_valid_stk]
    
def cal_t0_position(last_position, new_position, valid_position, valid_stk):
    assert np.sum(new_position<0)==0
    last_position = exclude_invalid(last_position, valid_stk)
    new_position = exclude_invalid(new_position, valid_stk)
    
    last_new_stk = list(set(last_position.index).union(new_position.index))
    t0_position = pd.DataFrame(index=last_new_stk, columns=['last', 'new', 'valid'], data=0)
    t0_position.loc[last_position.index, 'last'] = last_position
    t0_position.loc[new_position.index, 'new'] = new_position
    assert len(set(last_position.index)-set(valid_position.index))==0
    t0_position.loc[last_position.index, 'valid'] = valid_position[last_position.index]
    t0_position = t0_position.min(axis=1)
    t0_position = t0_position[t0_position>=100]
    t0_position.sort_index(inplace=True)
    return t0_position

def generate_t0_xlsx(t0_position, pf_code, today, time):
    assert np.sum(t0_position<=0)==0
    t0_position.index.name = '证券代码'
    t0_xlsx = pd.DataFrame(index=t0_position.index, 
                           columns=['买入交易账户', '卖出交易账户', '证券额度'], 
                           data=np.nan)

    t0_xlsx['买入交易账户'] = int(pf_code)
    t0_xlsx['卖出交易账户'] = int(pf_code)
    if pf_code=='5160501':
        t0_xlsx['买入交易账户'] = pf_code + 'l'
        t0_xlsx['卖出交易账户'] = pf_code + 'l'
    t0_xlsx['证券额度'] = t0_position
    
    t0_xlsx.to_excel(config_path.generate_pf_data_path+'t0_position/' + pf_code + '_日内可用额度_' + today + '_' + time + '.xlsx')
    
    return t0_xlsx

def cal_reb_position(last_position, new_position, valid_position, valid_stk):
    assert np.sum(new_position<0)==0
#    last_position = exclude_invalid(last_position, valid_stk) #@
#    new_position = exclude_invalid(new_position, valid_stk) #@
    
    last_new_stk = list(set(last_position.index).union(new_position.index))
    reb_position = pd.Series(index=last_new_stk, data=0)
    reb_position[new_position.index] = new_position
    reb_position[last_position.index] -= last_position
    sell_stk = reb_position[reb_position<0].index
    reb_position[sell_stk] = np.maximum(reb_position[sell_stk], -valid_position[sell_stk])
    reb_position = (reb_position / 100).astype(int) * 100
    reb_position = reb_position[reb_position!=0]
    reb_position.sort_index(inplace=True)
    return reb_position

def generate_reb_xlsx(reb_position, pf_code, today, time):
    assert np.sum(reb_position==0)==0
    reb_position.index.name = '证券代码'

    buy_stk = reb_position[reb_position>0].index.tolist()
    #@
    valid_stk = np.load(config_path.valid_stock_path + 'valid_stock_' + \
    today + '_' + time[:4] + '.npy').item()['trade_stk']
    buy_stk = sorted(list(set(buy_stk)&set(valid_stk)))
    print('generate_reb_xlsx',today,time,set(buy_stk)-set(valid_stk))
    
    assert len(set(buy_stk)-set(valid_stk))==0,print('buy invalid stk:',set(buy_stk)-set(valid_stk))
    sell_stk = reb_position[reb_position<0].index.tolist()
    stk_valid = sorted(list(set(sell_stk + buy_stk)))
    reb_position = reb_position.loc[stk_valid]

    reb_xlsx = pd.DataFrame(index=reb_position.index, 
                            columns=['委托方向', '指令数量', '指令价格', '价格模式', '市场代码'], 
                            data=np.nan)    
            
    reb_xlsx.loc[buy_stk, '委托方向'] = 1
    reb_xlsx.loc[sell_stk, '委托方向'] = 2
    assert np.sum(np.isnan(reb_xlsx['委托方向']))==0
    reb_xlsx['指令数量'] = reb_position.abs()
    reb_xlsx['指令价格'] = 0
    reb_xlsx['价格模式'] = 4
    sh_stk = [i for i in reb_position.index if i[-2:]=='SH']
    sz_stk = [i for i in reb_position.index if i[-2:]=='SZ']
    reb_xlsx.loc[sh_stk, '市场代码'] = 1
    reb_xlsx.loc[sz_stk, '市场代码'] = 2
    reb_xlsx = reb_xlsx.astype(int)
    
    reb_xlsx.to_excel(config_path.generate_pf_data_path+'reb_position/' + pf_code + '_下单指令_' + today + '_' + time + '.xlsx')
    
    return reb_xlsx

def generate_close_xlsx(reb_position, pf_code, today, close_after_open):
    assert np.sum(reb_position==0)==0
    reb_position.index.name = '证券代码'
    reb_xlsx = pd.DataFrame(index=reb_position.index, 
                            columns=['委托方向', '指令数量', '指令价格', '价格模式', '市场代码'], 
                            data=np.nan)
    reb_xlsx['委托方向'] = 2
    assert np.sum(np.isnan(reb_xlsx['委托方向']))==0
    reb_xlsx['指令数量'] = reb_position.abs()
    reb_xlsx['指令价格'] = 0
    reb_xlsx['价格模式'] = 4
    sh_stk = [i for i in reb_position.index if i[-2:]=='SH']
    sz_stk = [i for i in reb_position.index if i[-2:]=='SZ']
    reb_xlsx.loc[sh_stk, '市场代码'] = 1
    reb_xlsx.loc[sz_stk, '市场代码'] = 2
    reb_xlsx.index = [i[:-3] for i in reb_xlsx.index]
    reb_xlsx = reb_xlsx.astype(int)
    
    after = ''
    if close_after_open:
        after = '盘中'
    reb_xlsx.to_excel(config_path.generate_pf_data_path+'close_position/' + pf_code + '_下单指令_' + after + '平仓_' + today + '.xls')
    
    return reb_xlsx

def format_weight(w, pf_code, today, time, open_position=False):
    print('0',w.sum())
    files = os.listdir(pool_stock_path)
    files.sort()
    pool_file = files[-1]
    print('pool_file:',pool_file)
    qpool = pd.read_excel(pool_stock_path+pool_file, sheet_name=0, header=0, squeeze=False, converters={'证券代码':str,'市场名称':str})    
    qpool = qpool[(qpool['市场名称']=='2').values | (qpool['市场名称']=='1').values]
    assert np.sum(w<0)==0
    assert len(qpool)==len(set(qpool['证券代码'])),print('Error! 量化池中有重复证券代码')
    w = (w[w>0] * 100) / w[w>0].sum()
    w.index = [i[:-3]  for i in w.index]
    
    # qpool = pd.read_excel(config_path.generate_pf_data_path+'others/quant_pool.xls', converters={'证券名称':str, '证券代码':str})
    qpool.index = qpool['证券代码']
    qpool.index.name = None
    
    ################### to delete ###################
    # w = w.drop(sorted(set(w.index)-set(qpool.index)))  
    # w = (w[w>0] * 100) / w[w>0].sum()
    ################### to delete ###################
    
    assert len(set(w.index)-set(qpool.index))==0
    qpool['权重'] = np.nan
    qpool.loc[w.index.tolist(), '权重'] = w
    qpool['权重'] = (qpool['权重']*100)/(qpool['权重']).sum()
    # qpool['市场名称'] = [1 if s[0]=='6' else 2 for s in qpool['证券代码']]
    
    formatted_w = qpool[pd.notnull(qpool['权重'])]
    formatted_w = formatted_w[['证券代码', '证券名称', '市场名称', '权重']]
    assert abs(formatted_w['权重'].sum()-100)<0.0001
    file_name = config_path.generate_pf_data_path+'reb_weight/' + pf_code + '_weight_' + today + '.xlsx'
    if open_position:
        file_name = config_path.generate_pf_data_path+'open_weight/' + pf_code + '_weight_open_' + today + '_' + time + '.xlsx'
    formatted_w.to_excel(file_name)
    return formatted_w

def pf_generator(today_act, time, last_position, valid_position, valid_stk, pf_code, turn_ad, hedge_index, capital, cash, \
    barra_limit_dict, industry_loose, dupl_industry, split_fin, amt_limit_,single_stock_max_weight_,asset=None,mode=1,benchmark='500',test_label=False):
    print('capital',capital)
    print('turn_ad',turn_ad)
    close = pd.read_pickle(config_path.basic_data_path + 'daily/pre_close.pkl')
    left_stk = last_position[last_position<100].index.tolist()
    rebalance_stk = last_position[last_position>=100].index.tolist()
#    assert len(set(rebalance_stk)-set(close.columns))==0 ,print('@ Invalid stock in holding',set(rebalance_stk)-set(close.columns))
    print('@ Invalid stock in holding',set(rebalance_stk)-set(close.columns))
    rebalance_stk = sorted(list(set(rebalance_stk)&set(close.columns)))
    
    assert np.sum(last_position[rebalance_stk]<100)==0
    
    today = today_act.index[0]
    today_dt = today.strftime('%Y%m%d')  
    result1 = factor_tool.tradingday(today_dt,-2)
    last_dt = result1[-2]      
    
#    close.fillna(method='ffill',inplace=True)
#    close.loc[today] = np.nan
    print('pf_generator',today)
    close = close.loc[today]
    print('close',close)
#    pd.to_pickle(close[rebalance_stk],'/data/user/012620/own/AlphaData2/test.pkl')
    assert np.sum(np.isnan(close[rebalance_stk].values))==0
    if time=='1300':
        close = pd.read_pickle(config_path.basic_data_path + 'minute/Close/' + today_dt + '.pkl').iloc[119]
    if asset is None:
        asset = cash + (last_position[rebalance_stk] * close[rebalance_stk]).sum()
    assert asset>0,print('Error! assert <=0:',asset)
    # last_weight = pd.Series(index=close.index, data=0.)
    # last_weight[rebalance_stk] = (last_position[rebalance_stk] * close[rebalance_stk]) / asset
    
    lower_limit = pd.Series(index=close.index, data=0.)
    #@
    # lower_limit[rebalance_stk] = last_weight[rebalance_stk] - valid_position[rebalance_stk] * close[rebalance_stk] / asset    
    # if time=='0930':
        # assert np.abs(lower_limit.sum())<0.000001, 'lower limit error'
    w_path = config_path.weight_path + 'benchmark_'+benchmark+'/'
    print('time',time,'w_path',w_path)
    if time=='0930':
        real_w_path = w_path + 'real_w_am.pkl'
    elif time=='1300':
        real_w_path = w_path + 'real_w_pm.pkl'
    elif time in ['0930vwap','0930vwap300']:
        real_w_path = w_path + 'real_w_vwap.pkl'
    elif time=='0930lf':
        real_w_path = w_path + 'real_w_lf.pkl'
    else:
        raise AssertionError('wrong time input, supported times : 0930 and 1300 and 0930vwap')


    if mode==1:
        #capital = 5.e7 # @
        
        if time=='0930':
            real_w_path = w_path + 'real_w_am.pkl'
            last_weight = pd.read_pickle(w_path + 'simulation_w_pm.pkl').loc[last_dt]
            lower_limit = pd.Series(index=close.index, data=0.)
        elif time=='1300':
            real_w_path = w_path + 'real_w_pm.pkl'
            last_last_weight = pd.read_pickle(w_path + 'simulation_w_pm.pkl').loc[last_dt]
            last_weight = pd.read_pickle(w_path + 'real_w_am.pkl').loc[today_dt]
            lower_limit = np.maximum(last_weight - last_last_weight, 0.)
        elif time in ['0930vwap','0930vwap300']:
            real_w_path = w_path + 'real_w_vwap.pkl'
            last_weight = pd.read_pickle(w_path + 'simulation_w_vwap.pkl').loc[last_dt]
            lower_limit = pd.Series(index=close.index, data=0.) #@
        elif time=='0930lf':
            real_w_path = w_path + 'real_w_lf.pkl'
            last_weight = pd.read_pickle(w_path + 'simulation_w_lf.pkl').loc[last_dt]
            lower_limit = pd.Series(index=close.index, data=0.)        
        else:
            raise AssertionError('wrong time input, supported times : 0930 and 1300 and 0930vwap')
#    print('lower_limit',lower_limit)
    if hedge_index == 'HS300':
        last_weight.loc['601688.SH']=0
        last_weight.loc['600919.SH']=0
    try:
        if time in ['0930','1300']:
            w = optInd.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                            dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                            lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                            single_stock_max_weight=single_stock_max_weight_,simulate_label=False)
            print(pf_code,' New opt')
        else:
            if time not in ['0930vwap300']:
                w = optLF.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                                dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                                lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                                single_stock_max_weight=single_stock_max_weight_,simulate_label=False)
                print(pf_code,' New optLF')   
            else:
                w = optLF300.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                                dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                                lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                                single_stock_max_weight=single_stock_max_weight_,simulate_label=False)                
                print('@',time,pf_code,' New optLF300')         
        # w = optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
        #                  dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
        #                 lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
        #                 single_stock_max_weight=single_stock_max_weight_)
        # print(time,turn_ad,amt_limit_,single_stock_max_weight_,industry_loose,barra_limit_dict)
        # print(pf_code,' Old opt')
    except:                
        q_pool_stk = pd.read_pickle(config_path.weight_path + 'pool_stock.pkl')
        stock_set = pd.read_pickle(config_path.basic_data_path + 'daily/close.pkl').columns.tolist()
        valid_stk = list(set(q_pool_stk)&set(stock_set))
        if time in ['0930','1300']:
            w = optInd.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                                dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                                lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                                single_stock_max_weight=single_stock_max_weight_,simulate_label=False)            
        else:
            if time not in ['0930vwap300']:
                w = optLF.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                                dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                                lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                                single_stock_max_weight=single_stock_max_weight_,simulate_label=False)
                print(pf_code,' New optLF')   
            else:
                w = optLF300.optimize_hf({time[:4]:today_act}, {time[:4]:turn_ad}, hedge_index, capital, barra_limit_dict=barra_limit_dict, industry_loose=industry_loose, 
                                dupl_industry=dupl_industry, split_fin=split_fin, prev_weights=last_weight, \
                                lower_limit=lower_limit, pool_valid_stocks=valid_stk,amt_limit=amt_limit_,\
                                single_stock_max_weight=single_stock_max_weight_,simulate_label=False)                
                print(time,pf_code,' New optLF 0304')         
                  
        print(time,turn_ad,amt_limit_,single_stock_max_weight_,industry_loose,barra_limit_dict)                      
        lm.sendMessage("@Warning! %s Update Portfolio contains invalid_stk" % (pf_code))


    today_dt = today.strftime(format='%Y%m%d')
    w = w[time[:4]].loc[today]
    assert np.sum(w<0)==0
    if np.sum(w[pd.isnull(close)]!=0)!=0:
        tmp = (w[pd.isnull(close)])[w[pd.isnull(close)]!=0]
        ss = list(tmp.index)
        ws = tmp.values.tolist()
        ss_mes = ''
        for i in ss:
            ss_mes+=i+','
        ws_mes = ''
        for i in ws:
            ws_mes+=str(i)+','
        lm.sendMessage('Warning: NaN price in weight:' + ss_mes + ws_mes)    
#    assert np.sum(w[pd.isnull(close)]!=0)==0
    assert np.sum(np.isinf(w))==0
    assert np.sum(np.isnan(w))==0


    if test_label is False:#mode==1:
        real_w = pd.read_pickle(real_w_path) if os.path.exists(real_w_path) else pd.DataFrame(index=[today], columns=close.index, data=0.)
        real_w.loc[today, w.index] = w
        real_w.to_pickle(real_w_path)

    #fw = format_weight(w, pf_code, today_dt, time)
    
    print('sum of weight:', w.sum())
    #print('sum of adjusted weight:', fw['权重'].sum(), '<= 100 ?', str(fw['权重'].sum()<=100))
    
    today_position = ((w * asset) / close).fillna(0).astype(int)
    left_stk = list(set(left_stk)&set(w.index.tolist()))
    print('@',left_stk)
    today_position[left_stk] += last_position[left_stk]
    assert np.sum(today_position<0)==0
    assert np.sum(np.isnan(today_position))==0
    assert np.sum(np.isinf(today_position))==0
    today_position = today_position[today_position>0]
    fw = None
    return today_position, w, fw, asset




# def reb_to_dupl_index(last_position, cash, valid_stk, today_weight, today_dt, pf_code):
    
    # left_stk = last_position[last_position<100].index.tolist()
    # rebalance_stk = last_position[last_position>=100].index.tolist()
    # #if pf_code=='5160504':
    # #    invalid_stk = ['000333.SZ', '603156.SH', '601066.SH', '601838.SH', '601828.SH', '603259.SH', '601138.SH', '002925.SZ', '601598.SH', '600901.SH']
    # #elif pf_code=='5160803':
    # #    invalid_stk = ['601598.SH']
    # invalid_stk = []
    # left_stk = sorted(set(left_stk) - set(invalid_stk))
    # rebalance_stk = sorted(set(rebalance_stk) - set(invalid_stk))
    # assert np.sum(last_position[rebalance_stk]<100)==0
    
    # today = pd.Timestamp(today_dt)
    # close = pd.read_pickle('/app/HTSCAlpha/AlphaDataCenter/Basic/daily/close.pkl')
    # close.loc[today] = np.nan
    # close = close.shift(1).loc[today]
    # assert np.sum(np.isnan(close[rebalance_stk].values))==0
    
    # asset = cash + (last_position[rebalance_stk] * close[rebalance_stk]).sum()
    # assert asset>0
    # last_weight = pd.Series(index=close.index, data=0.)
    # last_weight[rebalance_stk] = (last_position[rebalance_stk] * close[rebalance_stk]) / asset

    # w = today_weight
    # assert np.sum(w<0)==0
    # assert np.sum(w[pd.isnull(close)]!=0)==0
    # assert np.sum(np.isinf(w))==0
    # assert np.sum(np.isnan(w))==0
    # fw = format_weight(w, pf_code, today_dt, time)
    
    # print('sum of weight:', w.sum())
    # print('sum of adjusted weight:', fw['权重'].sum(), '<= 100 ?', str(fw['权重'].sum()<=100))

    # today_position = ((w * asset) / close).fillna(0).astype(int)

    # today_position[left_stk] += last_position[left_stk]
    # assert np.sum(today_position<0)==0
    # assert np.sum(np.isnan(today_position))==0
    # assert np.sum(np.isinf(today_position))==0
    # today_position = today_position[today_position>0]
    
    # return today_position, w, fw



